from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook
import os

# Se crea la aplicación Flask
app = Flask(__name__)

# Usuario y contraseña para ingresar al sistema
USUARIO_CORRECTO = "admin"
CONTRASENA_CORRECTA = "1234"

# Lista donde se guardan temporalmente los contactos
contactos = []

# Nombre del archivo Excel donde se almacenan los datos
ARCHIVO = "contactos.xlsx"


# Función para crear el archivo Excel si no existe
def inicializar_excel():
    if not os.path.exists(ARCHIVO) or os.path.getsize(ARCHIVO) == 0:
        wb = Workbook()
        ws = wb.active

        # Encabezados de las columnas
        ws.append([
            "Nombre",
            "Teléfono",
            "Correo",
            "Última Visita",
            "Servicio Preferido",
            "Notas",
            "categoria",
            "Favorito"
        ])

        wb.save(ARCHIVO)

# Función para guardar un contacto en Excel
def guardar_en_excel(contacto):
    if not os.path.exists(ARCHIVO):
        inicializar_excel()

    try:
        wb = load_workbook(ARCHIVO)
        ws = wb.active

        # Se agrega una nueva fila con los datos del contacto
        ws.append([
            contacto["nombre"],
            contacto["telefono"],
            contacto["correo"],
            contacto["ultima_visita"],
            contacto["servicio"],
            contacto["notas"],
            contacto["categoria"],
            contacto["favorito"]
        ])

        wb.save(ARCHIVO)

    except Exception as e:
        print("Error al guardar en Excel:", e)
# Página principal
@app.route('/')
def index():
    return render_template('index.html')


# Inicio de sesión
@app.route('/login', methods=['GET', 'POST'])
def login():

    # Verifica si el usuario envió el formulario
    if request.method == 'POST':

        usuario = request.form['usuario']
        contrasena = request.form['contrasena']

        # Comprueba si los datos son correctos
        if usuario == USUARIO_CORRECTO and contrasena == CONTRASENA_CORRECTA:
            return redirect(url_for('lista'))

        # Muestra mensaje de error si los datos son incorrectos
        return render_template('login.html',
                               error='Usuario o contraseña incorrectos')

    return render_template('login.html')   # Muestra archivos html


# Lista de contactos
@app.route('/lista', methods=['GET', 'POST'])
def lista():

    # Se crea una copia de la lista de contactos
    resultados = contactos[:]

    # Buscar contacto por nombre
    if request.method == 'POST':
        nombre_buscar = request.form.get('nombre', '').strip().lower()

        if nombre_buscar:
            resultados = [
                c for c in contactos
                if nombre_buscar in c['nombre'].lower()
            ] #busca coincidencias

    # Ordenar contactos alfabéticamente
    ordenar = request.args.get('ordenar')

    if ordenar == 'nombre':
        resultados = sorted(
            resultados,
            key=lambda c: c['nombre'].lower()
        )

    return render_template('Lista.html', contactos=resultados)


# Agregar nuevo contacto
@app.route('/agregar', methods=['GET', 'POST'])
def agregar():

    if request.method == 'POST':

        # Obtener datos del formulario
        nombre = request.form['nombre']
        telefono = request.form.get('telefono', '')
        correo = request.form.get('correo', '')
        ultima_visita = request.form.get('ultima_visita', '')
        servicio = request.form.get('servicio', '')
        notas = request.form.get('notas', '')
        categoria = request.form.get('categoria', '')
        favorito = request.form.get('favorito', 'No')

        # Crear el nuevo contacto
        nuevo_contacto = {
            "id": len(contactos) + 1,
            "nombre": nombre,
            "telefono": telefono,
            "correo": correo,
            "ultima_visita": ultima_visita,
            "servicio": servicio,
            "notas": notas,
            "categoria": categoria,
            "favorito": favorito
        }

        # Guardar en la lista
        contactos.append(nuevo_contacto)

        # Guardar en Excel
        guardar_en_excel(nuevo_contacto)

        return redirect(url_for('lista'))

    return render_template('agregar.html')


# Mostrar la información completa de un contacto
@app.route('/detalle/<int:id>')
def detalle(id):

    contacto = next(
        (c for c in contactos if c['id'] == id),
        None
    )

    if contacto:
        return render_template('detalle.html', contacto=contacto)

    return "Contacto no encontrado", 404


# Editar un contacto existente
@app.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar(id):

    contacto = next(
        (c for c in contactos if c['id'] == id),
        None
    )

    if not contacto:
        return "Contacto no encontrado", 404

    if request.method == 'POST':

        # Actualizar los datos del contacto
        contacto['nombre'] = request.form['nombre']
        contacto['telefono'] = request.form.get('telefono', '')
        contacto['correo'] = request.form.get('correo', '')
        contacto['ultima_visita'] = request.form.get('ultima_visita', '')
        contacto['servicio'] = request.form.get('servicio', '')
        contacto['notas'] = request.form.get('notas', '')
        contacto['categoria'] = request.form.get('categoria', '')
        contacto['favorito'] = request.form.get('favorito', 'No')

        return redirect(url_for('lista'))

    return render_template('editar.html', contacto=contacto)


# Eliminar un contacto
@app.route('/eliminar/<int:id>', methods=['POST'])
def eliminar(id):

    global contactos

    contactos = [
        c for c in contactos
        if c['id'] != id
    ]

    return redirect(url_for('lista'))


# Generar reporte general
@app.route('/reporte')
def reporte():

    # Total de contactos registrados
    total_contactos = len(contactos)

    # Cantidad de contactos marcados como favoritos
    favoritos = sum(
        1 for c in contactos
        if c.get('favorito') == 'Sí'
    )

    return render_template(
        'reporte.html',
        total_contactos=total_contactos,
        favoritos=favoritos
    )


# Inicia la aplicación y crea el Excel si no existe
if __name__ == '__main__':
    inicializar_excel()
    app.run(debug=True)